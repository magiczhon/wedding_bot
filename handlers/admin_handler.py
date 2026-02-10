import os
import time
import json
import openpyxl
import telebot

from openpyxl.styles import Font, Alignment

from utils.logger import logger
from database.db import get_all_preferences, get_all_chat_ids, get_chat_ids


def register_handlers(bot: telebot.TeleBot):
    def get_admin_ids():
        with open('config.json', 'r') as f:
            config = json.load(f)
            return config.get("admin_ids", [])

    @bot.message_handler(func=lambda message: message.text == '📋 Выгрузка списка гостей')
    def view_guests(message):
        if message.from_user.id not in get_admin_ids():
            bot.send_message(message.chat.id, "У вас нет прав для выполнения этой операции.")
            return
            
        guests = get_all_preferences()
        logger.info(f'Получены данные гостей: {guests}')
        if not guests:
            bot.send_message(message.chat.id, "База данных пуста.")
            return
        
        # Создаем и отправляем Excel-файл
        filename = create_excel_file(guests)
        with open(filename, 'rb') as file:
            bot.send_document(message.chat.id, file)
        
        # Удаляем временный файл
        os.remove(filename)

    def create_excel_file(guests):
        logger.info('Create excel file with info about all guests')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Гости" # type: ignore
        
        # Заголовки
        headers = ["ФИО (регистрация)", "Telegram", "ФИО гостя", "Еда", "Напиток", "Дата обновления"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num) # type: ignore
            cell.value = header # type: ignore
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for row_idx, guest in enumerate(guests, 2):
            ws.cell(row=row_idx, column=1, value=guest[0]) # type: ignore
            ws.cell(row=row_idx, column=2, value=f"https://t.me/{guest[1]}" if guest[1] else "не указан") # pyright: ignore[reportOptionalMemberAccess]
            ws.cell(row=row_idx, column=3, value=guest[2]) # type: ignore
            ws.cell(row=row_idx, column=4, value=guest[3] if guest[3] else "не указано")  # type: ignore
            ws.cell(row=row_idx, column=5, value=guest[4] if guest[4] else "не указано")  # type: ignore # Напиток
            ws.cell(row=row_idx, column=6, value=str(guest[5]))  # Дата обновления # type: ignore
        
        # Автоширина столбцов
        for column in ws.columns: # type: ignore
            max_length = 0
            column_letter = column[0].column_letter # type: ignore
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width # type: ignore
        
        filename = "guests.xlsx"
        wb.save(filename)
        return filename

    @bot.message_handler(func=lambda message: message.text == '📨 Отправить сообщение всем пользователям бота')
    def request_broadcast_message(message):
        if message.from_user.id not in get_admin_ids():
            bot.send_message(message.chat.id, "У вас нет прав для выполнения этой операции.")
            return
        
        logger.info('Send message all users')
        bot.send_message(message.chat.id, "Введите сообщение, которое хотите разослать всем зарегистрированным пользователям (ВНИМАНИЕ! Если сообщение содержит более 4096 символов, отправь его тестовому пользователю):")
        bot.register_next_step_handler(message, process_broadcast_message)

    def process_broadcast_message(message):
        if message.from_user.id not in get_admin_ids():
            bot.send_message(message.chat.id, "У вас нет прав для выполнения этой операции.")
            return
        # Warning: Do not send more than about 4096 characters each message, otherwise you'll risk an HTTP 414 error. If you must send more than 4096 characters, use the split_string or smart_split function in util.py.
        broadcast_text = message.text
        chat_ids = get_all_chat_ids()
        
        sent_count = 0
        logger.info(f'Идет рассылка сообщения...')
        bot.send_message(message.chat.id, f'Идет рассылка сообщения...')

        for chat_id in chat_ids:
            try:
                bot.send_message(chat_id, broadcast_text)
                sent_count += 1
                time.sleep(0.5)
            except Exception as e:
                logger.error(f"Не удалось отправить сообщение в чат {chat_id}: {str(e)}")
        
        logger.info(f'Сообщение успешно отправлено {sent_count} пользователям')
        bot.send_message(message.chat.id, f"Рассылка завершена! Сообщение успешно отправлено {sent_count} пользователям.")


    @bot.message_handler(func=lambda message: message.text == 'Отправить сообщение ТЕСТОВЫМ пользователям')
    def request_broadcast_message_to_test_users(message):
        if message.from_user.id not in get_admin_ids():
            bot.send_message(message.chat.id, "У вас нет прав для выполнения этой операции.")
            return
        
        logger.info('Send message test users')
        bot.send_message(message.chat.id, "Введите сообщение, которое хотите отправить ТЕСТОВЫМ пользователям:")
        bot.register_next_step_handler(message, process_broadcast_message_to_test_users)

    def process_broadcast_message_to_test_users(message):
        if message.from_user.id not in get_admin_ids():
            bot.send_message(message.chat.id, "У вас нет прав для выполнения этой операции.")
            return
        # Warning: Do not send more than about 4096 characters each message, otherwise you'll risk an HTTP 414 error. If you must send more than 4096 characters, use the split_string or smart_split function in util.py.
        broadcast_text = message.text
        config = json.load(open('config.json', 'r'))
        user_ids = config['test_users_id']
        chat_ids = get_chat_ids(user_ids)
        
        for chat_id in chat_ids:
            try:
                logger.info(f'Высылаю сообщение в чат {chat_id = }')
                bot.send_message(chat_id, broadcast_text)
                time.sleep(0.5)
            except Exception as e:
                logger.error(f"Не удалось отправить сообщение в чат {chat_id}: {str(e)}")
        
        logger.info(f'Сообщение успешно отправлено пользователям {user_ids = }')
        bot.send_message(message.chat.id, f"Сообщение успешно отправлено пользователям {user_ids = }")